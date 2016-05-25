import csv
import numpy
from openpyxl import Workbook
from rmgpy.chemkin import getSpeciesIdentifier

"""
Assume ckcsv contains only one Soln
"""

def getROPFromCKCSV(ckcsvFile):
    """
    from ckcsv file get three dicts: firstColDict (e.g., Time/Distance_units:numpy.array)
    spc_total_dict (e.g., species_string: (header,numpy.array)), where header is sth like
    'spc ROP GasRxn Total_units', and 
    spc_indiv_dict (e.g., species_string: list of (header,numpy.array)), where header is 
    sth like 'spc ROP GasRxn#123_units'
    """
    firstColDict = {} # store Time/Distance series
    spc_total_dict = {}
    spc_indiv_dict = {}
    read_flag = False
    with open(ckcsvFile, 'r') as stream:
        reader = csv.reader(stream)
        for row in reader:
            if row[1].strip().startswith('rate-of-production'):
                read_flag = True
            if not read_flag:
                continue
            label = row[0].strip()
            tokens = label.split('_')
            if tokens[0] == 'Time':
                if 'Soln' in tokens[-1]:
                    raise Exception("This function only supports ckcsv with one Soln!")
                units = row[1].strip()[1:-1].lower()
                header = tokens[0] + '_(' + units + ')' 

                contentCol = numpy.array([float(r) for r in row[2:]], numpy.float)                
                contentCol *= {'sec': 1.0, 'min': 60., 'hr': 3600., 'msec': 1e-3, 'microsec': 1e-6}[units]
                firstColDict[header] = contentCol
                continue
            
            if tokens[0] == 'Distance':      
                if 'Soln' in tokens[-1]:
                    raise Exception("This function only supports ckcsv with one Soln!")

                units = row[1].strip()[1:-1].lower()
                header = tokens[0] + '_(' + units + ')'
                
                contentCol = numpy.array([float(r) for r in row[2:]], numpy.float)
                contentCol *= {'cm': 1.0, 'mm': 0.1, 'm': 100.}[units]
                firstColDict[header] = contentCol
                continue

            if len(tokens) > 1:
                if tokens[1] == 'ROP':
                    if 'Soln' in tokens[-1]:
                        raise Exception("This function only supports ckcsv with one Soln!")
                    species_string = tokens[0]
                    units = row[1].strip()[1:-1].lower()
                    header = ''
                    contentCol = numpy.array([float(r) for r in row[2:]], numpy.float)
                    if tokens[3] == 'Total':
                    	header += species_string + ' ROP ' + tokens[2] \
                    	+ ' ' + tokens[3] + '_(' + units + ')'
                        if species_string not in spc_total_dict:
                            spc_total_dict[species_string] = (header, contentCol)
                        else:
                            raise Exception("ckcsv file has two {} which is not in proper format!".format(header))
                    else: # where tokens[3] is something like GasRxn#123
                    	header += species_string + ' ROP ' \
                    	+ tokens[3] + '_(' + units + ')'
                        if species_string not in spc_indiv_dict:
                            spc_indiv_dict[species_string] = [(header, contentCol)]
                        else:
                            spc_indiv_dict[species_string].append((header, contentCol))


    return firstColDict, spc_total_dict, spc_indiv_dict

def getConcentrationDictFromCKCSV(ckcsvFile):
    """
    from ckcsv file get two dicts: firstColDict (e.g., Time/Distance_units:numpy.array)
    spc_conc_dict (e.g., species_string: numpy.array)
    """
    firstColDict = {} # store Time/Distance series
    spc_conc_dict = {}

    with open(ckcsvFile, 'r') as stream:
        reader = csv.reader(stream)
        for row in reader:
            label = row[0].strip()
            tokens = label.split('_')
            if tokens[0] == 'Time':
                if 'Soln' in tokens[-1]:
                    raise Exception("This function only supports ckcsv with one Soln!")
                units = row[1].strip()[1:-1].lower()
                header = tokens[0] + '_(' + units + ')' 

                contentCol = numpy.array([float(r) for r in row[2:]], numpy.float)                
                contentCol *= {'sec': 1.0, 'min': 60., 'hr': 3600., 'msec': 1e-3, 'microsec': 1e-6}[units]
                firstColDict[header] = contentCol
                continue
            
            if tokens[0] == 'Distance':      
                if 'Soln' in tokens[-1]:
                    raise Exception("This function only supports ckcsv with one Soln!")

                units = row[1].strip()[1:-1].lower()
                header = tokens[0] + '_(' + units + ')'
                
                contentCol = numpy.array([float(r) for r in row[2:]], numpy.float)
                contentCol *= {'cm': 1.0, 'mm': 0.1, 'm': 100.}[units]
                firstColDict[header] = contentCol
                continue
            # read concentration (mole fraction profile)
            if len(tokens) > 1:
                if tokens[0] == 'Mole' and tokens[1] == 'fraction':
                    if 'Soln' in tokens[-1]:
                        raise Exception("This function only supports ckcsv with one Soln!")
                    species_string = tokens[2]
                    contentCol = numpy.array([float(r) for r in row[2:]], numpy.float)
                    header = species_string + ' Mole_fraction'
                    if species_string not in spc_conc_dict:
                        spc_conc_dict[species_string] = contentCol
                    else:
                        raise Exception("ckcsv file has two {} which is not in proper format!".format(header))

    return firstColDict, spc_conc_dict

def saveROPXlsxFromCKCSV(ckcsvFile, xlsx_file, top_num):
    # write it into xlsx file
    firstColDict, spc_total_dict, spc_indiv_dict = getROPFromCKCSV(ckcsvFile)
    for species_string in spc_indiv_dict:
        spc_rxn_flux_sorted = sorted(spc_indiv_dict[species_string], \
            key=lambda tup: -abs(max(tup[1].min(), tup[1].max(), key=abs)))
        spc_indiv_dict[species_string] = spc_rxn_flux_sorted[:top_num]
    wb = Workbook()
    ws = wb.active
    for i, header in enumerate(firstColDict):
        ws.cell(row=1, column=i+1).value = header
        for j, cellValue in enumerate(firstColDict[header]):
            ws.cell(row=j+2, column=i+1).value = cellValue
    
    column_idx = len(firstColDict)
    for species_string in spc_total_dict:
        ws.cell(row=1, column=column_idx+1).value = spc_total_dict[species_string][0]
        for j, cellValue in enumerate(spc_total_dict[species_string][1]):
            ws.cell(row=j+2, column=column_idx+1).value = cellValue
        column_idx += 1
        try:
            for tup in spc_indiv_dict[species_string]:
                ws.cell(row=1, column=column_idx+1).value = tup[0]
                for j, cellValue in enumerate(tup[1]):
                    ws.cell(row=j+2, column=column_idx+1).value = cellValue
                column_idx += 1
        except KeyError:
            print "{} does not exist in spc_indiv_dict!".format(species_string)
    wb.save(xlsx_file)

def getFluxGraphEdgesDict(spc_rop_dict, core_reactions):
    
    graph_edges_dict = {}
    for rxn in core_reactions:
        for pair in rxn.pairs:
            if pair in graph_edges_dict:
                # get flux from spc_rop_dict
                species_string = getSpeciesIdentifier(pair[0])
                flux = getROPFlux(spc_rop_dict, species_string, rxn.index)
                if len(flux) > 0:
                    graph_edges_dict[pair][rxn] = flux
            elif (pair[1], pair[0]) in graph_edges_dict:
                # get flux from spc_rop_dict
                species_string = getSpeciesIdentifier(pair[1])
                flux = getROPFlux(spc_rop_dict, species_string, rxn.index)
                if len(flux) > 0:
                    graph_edges_dict[(pair[1], pair[0])][rxn] = flux
            else:
                # get flux from spc_rop_dict
                graph_edges_dict[pair] = {}
                species_string = getSpeciesIdentifier(pair[0])
                flux = getROPFlux(spc_rop_dict, species_string, rxn.index)
                if len(flux) > 0:
                    graph_edges_dict[pair][rxn] = flux
    return graph_edges_dict

def getROPFlux(spc_rop_dict, species_string, rxn_index):

    flux_tup_list = spc_rop_dict[species_string]
    for flux_tup in flux_tup_list:
        header = flux_tup[0]
        rxnNum = int(header.split("#")[1].split('_')[0])
        if rxnNum == rxn_index:
            flux = flux_tup[1]
            return flux
    return []